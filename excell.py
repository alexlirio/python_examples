#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import Workbook
 
 
if __name__ == '__main__':
 
    wb = Workbook()
     
    # grab the active worksheet
    ws = wb.active
     
    # Data can be assigned directly to cells
    ws['A1'] = 42
     
    # Rows can also be appended
    ws.append([1, 2, 3])
     
    # Python types will automatically be converted
    import datetime
    ws['A2'] = datetime.datetime.now()
     
    # Save the file
    wb.save("file/sample.xlsx")
    
    print 'Finished!'

